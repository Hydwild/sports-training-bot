"""Экспорт списка участников в Excel (openpyxl) и PDF (reportlab)."""
from __future__ import annotations

import io


def build_xlsx(title: str, when: str, location: str, max_p: int,
               rows: list[dict]) -> bytes:
    """rows: [{n, status, name, platform, attended, paid}]."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Участники"

    ws["A1"] = "Тренировка"; ws["B1"] = title
    ws["A2"] = "Дата"; ws["B2"] = when
    ws["A3"] = "Место"; ws["B3"] = location or "-"
    ws["A4"] = "Лимит"; ws["B4"] = max_p
    for r in range(1, 5):
        ws[f"A{r}"].font = Font(bold=True)

    header_row = 6
    headers = ["№", "Статус", "Имя", "Платформа", "Пришёл", "Оплатил"]
    fill = PatternFill("solid", fgColor="3A7BD5")
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=header_row, column=col, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = fill

    for i, row in enumerate(rows, 1):
        ws.cell(row=header_row + i, column=1, value=row["n"])
        ws.cell(row=header_row + i, column=2, value=row["status"])
        ws.cell(row=header_row + i, column=3, value=row["name"])
        ws.cell(row=header_row + i, column=4, value=row["platform"])
        ws.cell(row=header_row + i, column=5, value=row["attended"])
        ws.cell(row=header_row + i, column=6, value=row["paid"])

    for col, width in zip("ABCDEF", (12, 12, 28, 12, 10, 10), strict=False):
        ws.column_dimensions[col].width = width

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def build_pdf(title: str, when: str, location: str, max_p: int,
              rows: list[dict]) -> bytes:
    """PDF со списком участников. Кириллица через встроенный шрифт."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import (
        Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, title=title)
    styles = getSampleStyleSheet()
    elements = [
        Paragraph(title, styles["Title"]),
        Paragraph(f"Дата: {when}", styles["Normal"]),
        Paragraph(f"Место: {location or '-'}", styles["Normal"]),
        Paragraph(f"Лимит: {max_p}", styles["Normal"]),
        Spacer(1, 12),
    ]
    data = [["№", "Статус", "Имя", "Платформа", "Пришёл", "Оплатил"]]
    for row in rows:
        data.append([row["n"], row["status"], row["name"], row["platform"],
                     row["attended"], row["paid"]])
    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#3A7BD5")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1),
         [colors.white, colors.HexColor("#F0F4FA")]),
    ]))
    elements.append(table)
    doc.build(elements)
    buf.seek(0)
    return buf.read()
